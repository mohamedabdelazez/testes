import tkinter as tk
from tkinter import ttk
import math
from fractions import Fraction
import openpyxl
from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
# List to store all calculations
calculation_data = []

# GUI construction
root = tk.Tk()
root.title("Water Pipe Diameter and Head Loss Calculator")
pipe_data_constant_speed_none_noise_sensitive = {
    "pipe_diameter_inches": [Fraction(1, 2), Fraction(3, 4), 1, 1.25, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10, 12, 14, 16, 18, 20, 24, 26],
    "flow_range_gpm_at_2000 hours": [(0.5, 5), (5, 12), (12, 19), (19, 34), (34, 57), (57, 73), (73, 100), (100, 180), (180, 320), (320, 430), (430, 700), (700, 1200), (1200, 1900), (1900, 2900), (2900, 4000), (4000, 4900), (4900, 7000), (7000, 7700), (7700, 12000), (12000, 14000)],
    "flow_range_gpm_at_4400 hours": [(0.5, 3.9), (4, 9), (9, 14), (14, 26), (26, 43), (43, 55), (55, 77), (77, 140), (140, 240), (240, 330), (330, 530), (530, 900), (900, 1500), (1500, 2200), (2200, 3000), (3000, 3800), (3800, 5300), (5300, 5800), (5800, 8900), (8900, 11000)],
    "flow_range_gpm_at_8760 hours": [(0.5, 3), (3, 7.0), (7, 11), (11, 20), (20, 34), (34, 44), (44, 60), (60, 110), (110, 190), (190, 260), (260, 420), (420, 720), (720, 1200), (1200, 1700), (1700, 2400), (2400, 3000), (3000, 4200), (4200, 4600), (4600, 7100), (7100, 8500)]
}


pipe_data_constant_speed_noise_sensitive = {
    "pipe_diameter_inches": [Fraction(1, 2), Fraction(3, 4), 1, 1.25, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10, 12, 14, 16, 18, 20, 24, 26],
    "flow_range_gpm_at_2000 hours": [(0.5, 1.8), (1.9, 4.6), (4.7, 8.9), (9, 15), (15.1, 24), (24.1, 51), (51.1, 81), (81.1, 140), (140.1, 280), (280.1, 430), (430.1, 700), (700.1, 1200), (1200.1, 1900), (1900.1, 2900), (2900.1, 4000), (4000.1, 4900), (4900.1, 7000), (7000.1, 7700), (7700.1, 12000), (12000.1, 14000)],
    "flow_range_gpm_at_4400 hours": [(0.5, 1.8), (1.9, 4.6), (4.7, 8.9), (9, 15), (15.1, 24), (24.1, 51), (51.1, 77), (77.1, 140), (140.1, 240), (240.1, 330), (330.1, 530), (530.1, 900), (900.1, 1500), (1500.1, 2200), (2200.1, 3000), (3000.1, 3800), (3800.1, 5300), (5300.1, 5800), (5800.1, 8900), (8900.1, 11000)],
    "flow_range_gpm_at_8760 hours": [(0.5, 1.8), (1.9, 4.6), (4.7, 8.9), (9, 15), (15.1, 24), (24.1, 44), (44.1, 60), (60.1, 110), (110.1, 190), (190.1, 260), (260.1, 420), (420.1, 720), (720.1, 1200), (1200.1, 1700), (1700.1, 2400), (2400.1, 3000), (3000.1, 4200), (4200.1, 4600), (4600.1, 7100), (7100.1, 8500)]
}

pipe_data_varible_speed_noise_sensitive = {
    "pipe_diameter_inches": [Fraction(1, 2), Fraction(3, 4), 1, 1.25, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10, 12, 14, 16, 18, 20, 24, 26],
    "flow_range_gpm_at_2000 hours": [(0.3, 1.8), (1.9, 4.6), (4.7, 8.9), (9, 15), (15.1, 24), (24.1, 51), (51.1, 81), (81.1, 140), (140.1, 280), (280.1, 490), (490.1, 770), (770.1, 1500), (1500.1, 2700), (2700.1, 4200), (4200.1, 5400), (5400.1, 7200), (7200.1, 9200), (9200.1, 11000), (11000.1, 17000), (17000.1, 20000)],
    "flow_range_gpm_at_4400 hours": [(0.3, 1.8), (1.9, 4.6), (4.7, 8.9), (9, 15), (15.1, 24), (24.1, 51), (51.1, 81), (81.1, 140), (140.1, 280), (280.1, 490), (490.1, 770), (770.1, 1400), (1400.1, 2200), (2200.1, 3300), (3300.1, 4600), (4600.1, 5700), (5700.1, 8000), (8000.1, 8800), (8800.1, 13000), (13000.1, 16000)],
    "flow_range_gpm_at_8760 hours": [(0.3, 1.8), (1.9, 4.6), (4.7, 8.9), (9, 15), (15.1, 24), (24.1, 51), (51.1, 81), (81.1, 140), (140.1, 280), (280.1, 390), (390.1, 630), (630.1, 1100), (1100.1, 1800), (1800.1, 2600), (2600.1, 3600), (3600.1, 4500), (4500.1, 6300), (6300.1, 7000), (7000.1, 11000), (11000.1, 13000)]
}
pipe_data_varible_speed_none_noise_sensitive = {
    "pipe_diameter_inches": [Fraction(1, 2), Fraction(3, 4), 1, 1.25, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10, 12, 14, 16, 18, 20, 24, 26],
    "flow_range_gpm_at_2000 hours": [(0.3, 7.8), (7.9, 18), (18.1, 29), (29.1, 51), (51.1, 88), (88.1, 120), (120.1, 160), (160.1, 270), (270.1, 480), (480.1, 670), (670.1, 1100), (1100.1, 1800), (1800.1, 2900), (2900.1, 4400), (4400.1, 6000), (6000.1, 7400), (7400.1,  10000), (10000.1, 11000), (11000.1, 17000), (17000.1, 21000)],
    "flow_range_gpm_at_4400 hours": [(0.2, 5.9), (6, 14), (14.1, 22), (22.1, 39), (39.1, 67), (67.1, 84), (84.1, 120), (120.1, 210), (210.1, 360), (360.1, 510), (510.1, 800), (800.1, 1400), (1400.1, 2200), (2200.1, 3300), (3300.1, 4600), (4600.1, 5700), (5700.1, 8000), (8000.1, 8800), (8800.1, 13000), (13000.1, 16000)],
    "flow_range_gpm_at_8760 hours": [(0.1, 4.6), (4.7, 11), (11.1, 17), (17.1, 30), (30.1, 52), (52.1, 67), (67.1, 91), (91.1, 160), (160.1, 290), (290.1, 390), (390.1, 630), (630.1, 1100), (1100.1, 1800), (1800.1, 2600), (2600.1, 3600), (3600.1, 4500), (4500.1, 6300), (6300.1, 7000), (7000.1, 11000), (11000.1, 13000)]
}

pipe_material_inner_diameter_inches = {
	"HDPE": [0.622, 0.824, 1.049, 1.38, 1.61, 2.067, 2.469, 3.068, 4.026, 5.047, 6.065, 7.981, 10.02, 11.938, 13.876, 15.812, 17.748, 19.624, 23.562, 25.376],
	"Black Steel schd 40": [0.622, 0.824, 1.049, 1.38, 1.61, 2.067, 2.469, 3.068, 4.026, 5.047, 6.065, 7.981, 10.02, 11.938, 13.876, 15.812, 17.748, 19.624, 23.562, 25.376],
	"Black Steel schd 80": [0.622, 0.824, 1.049, 1.38, 1.61, 2.067, 2.469, 3.068, 4.026, 5.047, 6.065, 7.981, 10.02, 11.938, 13.876, 15.812, 17.748, 19.624, 23.562, 25.376],
	"PVC": [0.622, 0.824, 1.049, 1.38, 1.61, 2.067, 2.469, 3.068, 4.026, 5.047, 6.065, 7.981, 10.02, 11.938, 13.876, 15.812, 17.748, 19.624, 23.562, 25.376],
	
}

pipe_material_outer_diameter_inches = {
	"HDPE": [0.622, 0.824, 1.049, 1.38, 1.61, 2.067, 2.469, 3.068, 4.026, 5.047, 6.065, 7.981, 10.02, 11.938, 13.876, 15.812, 17.748, 19.624, 23.562, 25.376],
	"Black Steel schd 40": [0.622, 0.824, 1.049, 1.38, 1.61, 2.067, 2.469, 3.068, 4.026, 5.047, 6.065, 7.981, 10.02, 11.938, 13.876, 15.812, 17.748, 19.624, 23.562, 25.376],
	"Black Steel schd 80": [0.622, 0.824, 1.049, 1.38, 1.61, 2.067, 2.469, 3.068, 4.026, 5.047, 6.065, 7.981, 10.02, 11.938, 13.876, 15.812, 17.748, 19.624, 23.562, 25.376],
	"PVC": [0.622, 0.824, 1.049, 1.38, 1.61, 2.067, 2.469, 3.068, 4.026, 5.047, 6.065, 7.981, 10.02, 11.938, 13.876, 15.812, 17.748, 19.624, 23.562, 25.376],
	
}


# Function to calculate velocity in ft/s using inner diameter
def calculate_velocity(flow_rate, inner_diameter):
    # Convert flow rate from gpm to ft^3/s
    flow_rate_ft3s = flow_rate / 448.831
    # Calculate velocity in ft/s
    velocity_fts = flow_rate_ft3s / (math.pi * (inner_diameter/12)**2 / 4)
    return velocity_fts

# Function to calculate head loss using inner diameter
def calculate_head_loss(flow_rate, inner_diameter, fluid_density, fluid_viscosity, pipe_length, pipe_material):
    # Convert inner diameter from inches to meters
    inner_diameter_mm = inner_diameter * 25.4
    inner_diameter_m = inner_diameter_mm / 1000
    # Convert flow rate from gpm to m^3/s
    flow_rate_ls = flow_rate * 0.06
    flow_rate_m3s = flow_rate_ls / 1000
    # Calculate pipe area
    pipe_area = (inner_diameter_m ** 2) * math.pi / 4
    # Calculate velocity
    velocity = flow_rate_m3s / pipe_area
    # Calculate Reynolds number
    reynolds_number = velocity * inner_diameter_m * fluid_density / fluid_viscosity
    # Get roughness based on pipe material
    pipe_material_and_roughness = {
        "Black Steel schd 40": 0.045 / 1000,
        "Black Steel schd 80": 0.045 / 1000,
        "HDPE": 0.0007 / 1000,
        "PVC": 5 / 1000
    }
    roughness = pipe_material_and_roughness.get(pipe_material, 0.045 / 1000)  # Default to black steel schd 40 if material not found
    # Calculate friction factor
    friction_factor = calculate_friction_factor(inner_diameter_m, reynolds_number, roughness)
    # Calculate head loss
    head_loss = (friction_factor * pipe_length * velocity ** 2) / (2 * 9.81 * inner_diameter_m)
    return head_loss

# Function to calculate friction factor using Colebrook-White equation
def calculate_friction_factor(diameter, reynolds, roughness):
    friction = 0.08  # Starting friction factor
    while True:
        left_f = 1 / friction ** 0.5
        right_f = -2 * math.log10((2.51 / (reynolds * friction ** 0.5)) + (roughness / (3.72 * diameter)))
        friction = friction - 0.000001  # Change friction factor
        if right_f - left_f <= 0:  # Check if left = right
            break
    return friction

# Create a checkbox for manual diameter selection
manual_diameter_var = tk.BooleanVar()
manual_diameter_checkbox = ttk.Checkbutton(root, text="Manual Diameter Selection", variable=manual_diameter_var, onvalue=True, offvalue=False)
manual_diameter_checkbox.grid(row=12, column=0, columnspan=1, padx=10, pady=10)

# Create a dropdown menu for pipe diameters
#diameter_label = ttk.Label(root, text="Choose Pipe Diameter:")
#diameter_label.grid(row=12, column=1, padx=10, pady=10)

diameter_var = tk.StringVar()
diameter_dropdown = ttk.Combobox(root, textvariable=diameter_var, values=[])
diameter_dropdown.grid(row=12, column=1, padx=10, pady=10)

from fractions import Fraction

def populate_diameter_dropdown():
  if manual_diameter_var.get():
    # Convert all values to string representations, including fractions
    diameters = [str(d) for d in pipe_data_constant_speed_none_noise_sensitive["pipe_diameter_inches"]]
  else:
    # Convert to float if not already a Fraction, otherwise keep as Fraction
    diameters = [str(get_diameter(d)) for d in pipe_data_constant_speed_none_noise_sensitive["pipe_diameter_inches"]]
  diameter_dropdown['values'] = diameters

def calculate_diameter():
    selected_system = system_var.get()
    selected_noise = noise_var.get()
    selected_hours = hours_var.get().strip() 
    
    # Check if the length entry is empty
    if length_entry.get() == "":
        messagebox.showerror("Error", "Please enter the pipe length.")
        return

    # Check if the flow rate entry is empty
    if flow_entry.get() == "":
        messagebox.showerror("Error", "Please enter the flow rate.")
        return

    if selected_system == "Constant Speed":
        if selected_noise == "Yes":
            pipe_data = pipe_data_constant_speed_noise_sensitive
        elif selected_noise == "No":
            pipe_data = pipe_data_constant_speed_none_noise_sensitive
    elif selected_system == "Variable Speed":
        if selected_noise == "Yes":
            pipe_data = pipe_data_varible_speed_noise_sensitive
        elif selected_noise == "No":
            pipe_data = pipe_data_varible_speed_none_noise_sensitive

    flow_rate = float(flow_entry.get())
    pipe_length = float(length_entry.get()) 
    fluid_density = 998  # Default density of water at 5C
    fluid_viscosity = 0.0015182  # Default viscosity of water at 5C (Pa.s)

    if manual_diameter_var.get():
        selected_diameter_str = diameter_var.get()  # Get the string representation of the fraction
        # Convert the selected diameter to fraction if necessary
        if selected_diameter_str == "1/2":
            selected_diameter = Fraction(1, 2)
        elif selected_diameter_str == "3/4":
            selected_diameter = Fraction(3, 4)
        else:
            selected_diameter = float(selected_diameter_str)
    else:
        # Find the appropriate diameter
        for min_flow, max_flow in pipe_data["flow_range_gpm_at_" + selected_hours]:
            if min_flow <= flow_rate <= max_flow:
                selected_diameter = pipe_data["pipe_diameter_inches"][pipe_data["flow_range_gpm_at_" + selected_hours].index((min_flow, max_flow))]
                break

    # Get inner diameter using the selected material
    material = material_var.get()
    inner_diameter = pipe_material_inner_diameter_inches[material][pipe_data["pipe_diameter_inches"].index(selected_diameter)]
    external_diameter = pipe_material_outer_diameter_inches[material][pipe_data["pipe_diameter_inches"].index(selected_diameter)]

    head_loss = calculate_head_loss(flow_rate, inner_diameter, fluid_density, fluid_viscosity, pipe_length, material)
    velocity = calculate_velocity(flow_rate, inner_diameter)

    internal_label.config(text=f"Internal Diameter: {inner_diameter} inches")
    external_label.config(text=f"External Diameter: {external_diameter} inches")
    result_label.config(text=f"Nominal Diameter : {selected_diameter} inches")
    head_loss_label.config(text=f"Head Loss: {head_loss:.2f} meters")
    velocity_label.config(text=f"Velocity: {velocity:.2f}ft/s")

    # Store calculation data
    calculation_data.append({
        "nominal_diameter": selected_diameter,
        "internal_diameter": inner_diameter,
        "external_diameter": external_diameter,
        "head_loss": head_loss,
        "velocity": velocity
    })

# Connect the checkbox to the dropdown menu
def update_diameter_dropdown(*args):
	
    if manual_diameter_var.get():
        populate_diameter_dropdown()
        diameter_dropdown.config(state="normal")  # Enable the dropdown
    else:
        diameter_dropdown.config(state="disabled")  # Disable the dropdown
        diameter_dropdown.set("")  # Clear the selection

# Set the diameter dropdown to disabled initially
diameter_dropdown.config(state="disabled")

manual_diameter_var.trace_add('write', update_diameter_dropdown)

# Label to display internal diameter
internal_label = ttk.Label(root, text="")
internal_label.grid(row=10, column=0, columnspan=2, padx=10, pady=10)

# Label to display external diameter
external_label = ttk.Label(root, text="")
external_label.grid(row=11, column=0, columnspan=2, padx=10, pady=10)

# Dropdown to choose pipe material
material_label = ttk.Label(root, text="Choose Pipe Material:")
material_label.grid(row=5, column=0, padx=10, pady=10)

material_var = tk.StringVar()
material_dropdown = ttk.Combobox(root, textvariable=material_var, values=["Black Steel schd 40", "Black Steel schd 80", "HDPE", "PVC" ])
material_dropdown.grid(row=5, column=1, padx=10, pady=10)
material_dropdown.current(0)  # Set default value

# Dropdown to choose system type (constant speed or variable speed)
system_label = ttk.Label(root, text="Choose System Type:")
system_label.grid(row=0, column=0, padx=10, pady=10)

system_var = tk.StringVar()
system_dropdown = ttk.Combobox(root, textvariable=system_var, values=["Constant Speed", "Variable Speed"])
system_dropdown.grid(row=0, column=1, padx=10, pady=10)
system_dropdown.current(0)  # Set default value

# Dropdown to choose if the application is noise sensitive or not
noise_label = ttk.Label(root, text="Is the Application Noise Sensitive?")
noise_label.grid(row=1, column=0, padx=10, pady=10)

noise_var = tk.StringVar()
noise_dropdown = ttk.Combobox(root, textvariable=noise_var, values=["Yes", "No"])
noise_dropdown.grid(row=1, column=1, padx=10, pady=10)
noise_dropdown.current(1)  # Set default value

# Dropdown to choose the number of operating hours per year
hours_label = ttk.Label(root, text="Choose Number of Operating Hours per Year:")
hours_label.grid(row=2, column=0, padx=10, pady=10)

hours_var = tk.StringVar()
hours_dropdown = ttk.Combobox(root, textvariable=hours_var, values=["2000 hours", "4400 hours", "8760 hours"])
hours_dropdown.grid(row=2, column=1, padx=10, pady=10)
hours_dropdown.current(0)  # Set default value

# Entry field to get flow rate
flow_label = ttk.Label(root, text="Enter Flow Rate (gpm):")
flow_label.grid(row=3, column=0, padx=10, pady=10)

flow_entry = ttk.Entry(root)
flow_entry.grid(row=3, column=1, padx=10, pady=10)

# Entry field to get pipe length
length_label = ttk.Label(root, text="Enter Pipe Length (m):")
length_label.grid(row=4, column=0, padx=10, pady=10)

length_entry = ttk.Entry(root)
length_entry.grid(row=4, column=1, padx=10, pady=10)

# Function to calculate when the button is clicked
calculate_button = ttk.Button(root, text="Calculate Diameter", command=calculate_diameter)
calculate_button.grid(row=6, column=0, columnspan=2, padx=10, pady=10)

# Text area to display calculation results
result_label = ttk.Label(root, text="")
result_label.grid(row=7, column=0, columnspan=2, padx=10, pady=10)

head_loss_label = ttk.Label(root, text="")
head_loss_label.grid(row=8, column=0, columnspan=2, padx=10, pady=10)

# Label to display velocity
velocity_label = ttk.Label(root, text="")
velocity_label.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

def export_to_excel():
    # Create a new Excel workbook
    wb = Workbook()
    
    # Sheet for pipe calculations
    sheet = wb.active
    sheet.title = "Pipe Calculations"

    # Write headers for calculations
    headers = ["Nominal Diameter (inches)", "Internal Diameter (inches)", "External Diameter (inches)", "Head Loss (meters)", "Velocity (ft/s)"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Write calculation data
    for i, data in enumerate(calculation_data, start=2):
        sheet.cell(row=i, column=1, value=str(data["nominal_diameter"])).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=i, column=2, value=data["internal_diameter"]).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=i, column=3, value=data["external_diameter"]).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=i, column=4, value=data["head_loss"]).alignment = Alignment(horizontal='center', vertical='center')
        sheet.cell(row=i, column=5, value=data["velocity"]).alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths for calculation sheet
    column_widths = [26, 25, 25, 20, 20]  # Adjust according to your preference
    for i, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = width

    # Sheet for total length
    total_length_sheet = wb.create_sheet(title="Total Length")

    # Write headers for total length
    total_length_headers = ["Nominal Diameter (inches)", "Total Length (meters)"]
    for col, header in enumerate(total_length_headers, start=1):
        cell = total_length_sheet.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Calculate total length for each pipe diameter
    total_lengths = {}
    for data in calculation_data:
        diameter = data["nominal_diameter"]
        length = float(length_entry.get())
        if diameter in total_lengths:
            total_lengths[diameter] += length
        else:
            total_lengths[diameter] = length

    # Write total lengths to the sheet
    for i, (diameter, length) in enumerate(total_lengths.items(), start=2):
        total_length_sheet.cell(row=i, column=1, value=str(diameter)).alignment = Alignment(horizontal='center', vertical='center')
        total_length_sheet.cell(row=i, column=2, value=length).alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths for total length sheet
    total_length_sheet.column_dimensions['A'].width = 25
    total_length_sheet.column_dimensions['B'].width = 20

    # Ask user to choose where to save the file
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Workbook", "*.xlsx")])
    if file_path:
        # Save the workbook
        wb.save(file_path)
        
        messagebox.showinfo("Export Complete", "Data exported to Excel successfully!")

# Button to export data to Excel
export_button = ttk.Button(root, text="Export to Excel", command=export_to_excel)
export_button.grid(row=13, column=0, columnspan=2, padx=10, pady=10)



root.mainloop()
